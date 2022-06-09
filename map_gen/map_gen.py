#python3读写excel

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl.styles import colors  # 导入填充模块
import openpyxl
import os
import re
import sys
import numpy as np

def size2dec(xx):    
    if(xx[-1]=='G'):
        yy = int(re.sub("\D", "", xx))*1024*1024*1024
    elif(xx[-1]=='M'):
        yy = int(re.sub("\D", "", xx))*1024*1024
    elif(xx[-1]=='K'):
        yy = int(re.sub("\D", "", xx))*1024
    else:
        yy = int(re.sub("\D", "", xx))    
    return yy


def write_cell(ws,x,y,value,c = None):
    ws.cell(x,y).value = value        
    if(c is not None):
        #print(c)
##        fill1 = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
##        font1 = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])  # 设置字体样式

        if(c[0] != 'ffffff'):            
            fill1 = PatternFill('solid', fgColor=c[0])  # 设置填充颜色为 橙色
            ws.cell(x,y).fill = fill1
        font1 = Font(bold=True, color=c[1])  # 设置字体样式        
        ws.cell(x,y).font = font1


def conv_write_str(value,tp = 'origin'):
    if(tp == 'hex'):
        if(value < 4*1024*1024*1024):
            value_hex = '{:08X}'.format(value)
            hex_str = '0x'+value_hex[0:4]+'_'+ value_hex[4:8]            
        else:
            value_hex = '{:010X}'.format(value)
            #hex_str = '0x'+value_hex[0:2]+'_'+value_hex[2:6]+'_'+ value_hex[6:10]
            hex_str = '0x'+value_hex[0:6]+'_'+ value_hex[6:10]
        return hex_str
    elif(tp == 'size'):
        tmp_size = value
        if(tmp_size < 1024):
            tmp_hex = int(tmp_size)            
        elif(tmp_size >= 1024 and tmp_size < 1024*1024):
            tmp_hex = str(int(tmp_size/1024))+'K'
        elif(tmp_size >= 1024*1024 and tmp_size < 1024*1024*1024):
            tmp_hex = str(int(tmp_size/1024/1024))+'M'
        else:
            tmp_hex = str(int(tmp_size/1024/1024/1024))+'G'
        return tmp_hex
    elif(tp == 'origin'):
        return value
    else:        
        print('error-----%s',tp)
        return 'error'
    
     



wb1=openpyxl.load_workbook('origin.xlsx',data_only = "true")
wb2=Workbook()
sheets1=wb1.sheetnames#获取sheet页
print(sheets1)
nsheets = len(sheets1)
style_warning= ['ff0000','ffff00']
style = ['87cdfa','000000']
style_insert = ['ffffff','ff0000']

#for list in sheets1:
for i in range(nsheets):
    print(sheets1[i])
    #sheet1 = wb1[list1] #index by name 
    ws_temp = wb1.worksheets[i] #sheet by index
    #wb_sheet = wb2.create_sheet(title=sheets1[i],index=i)
    
    num_rows = ws_temp.max_row
    num_cols = ws_temp.max_column
    print(num_rows)
    print(num_cols)
    sta_ind = -1
    end_ind = -1
    siz_ind = -1
    vsz_ind = -1

    for j in range(1,num_cols+1):
        if ws_temp.cell(1,j).value == 'start address':
            sta_ind = j
        elif ws_temp.cell(1,j).value == 'end address':
            end_ind = j
        elif ws_temp.cell(1,j).value == 'size':
            siz_ind = j
        elif ws_temp.cell(1,j).value == 'virtual size':
            vsz_ind = j
    if sta_ind>=0 and end_ind>=0 and siz_ind >= 0 and vsz_ind>=0:
        print(sheets1[i]+'  start:'+str(sta_ind)+'  end:'+str(end_ind)+'  size:'+str(siz_ind)+'  virtual size:'+str(vsz_ind))
        wb_sheet = wb2.create_sheet(title=sheets1[i],index=i)

        ins_flag = np.zeros(num_rows+1,dtype=int)
        
        for m in range(1,num_rows+1):
            for n in range(1,num_cols+1):
                write_cell(wb_sheet,m,n,ws_temp.cell(m,n).value)
                font_cp = Font(color=ws_temp.cell(m,n).font.color)  # 设置字体样式 
                #print(ws_temp.cell(m,n).font)
                wb_sheet.cell(m,n).font = font_cp

        for k in range(2,num_rows+1):
            #print(k)
            if(k==2):
                sta_last = ''
                vsiz_dec_last = 0
            else:
                sta_last = sta_value
                vsiz_dec_last = vsize_dec
                                
            sta_value = ws_temp.cell(k,sta_ind).value        
            end_value = ws_temp.cell(k,end_ind).value
            siz_value = ws_temp.cell(k,siz_ind).value
            vsz_value = ws_temp.cell(k,vsz_ind).value


            if(k<num_rows):
                sta_next = ws_temp.cell(k+1,sta_ind).value
            else:
                sta_next = ''


            if(sta_value == 15 or sta_value ==23 or sta_value is None):
                sta_value = ''
                
            if(end_value == 15 or end_value ==23 or end_value is None):
                end_value = ''                    

            if(siz_value == 15 or siz_value ==23 or siz_value is None):
                siz_value = ''

            if(vsz_value == 15 or vsz_value ==23 or vsz_value is None):
                vsz_value = ''

            if(sta_next == 15 or sta_next ==23 or sta_next is None):
                sta_next = ''                    
            
            if(siz_value != ''):                
                size_dec = size2dec(siz_value)                
            else:
                size_dec = 0
                
            if(vsz_value != ''):                
                vsize_dec = size2dec(vsz_value)
            else:
                vsize_dec = 0

            if(sta_value == ''):
                if(sta_next != '' and vsize_dec > 0):
                    tmp_start = int(sta_next,16)-vsize_dec;
                    sta_value = conv_write_str(tmp_start,'hex')
                    write_cell(wb_sheet,k,sta_ind,sta_value,style)
                    
                elif(sta_last != '' and vsiz_dec_last >0):
                    tmp_start = int(sta_last,16)+vsiz_dec_last;
                    sta_value = conv_write_str(tmp_start,'hex')
                    write_cell(wb_sheet,k,sta_ind,conv_write_str(tmp_start,'hex'),style)                   
                else:
                    write_cell(wb_sheet,k,sta_ind,'N/A!',style_warning)
                    sta_value = ''

            if(sta_value != '' and sta_last != ''):
                tmp_size = int(sta_value,16) - int(sta_last,16)
                if(tmp_size > vsiz_dec_last):
                    ins_flag[k] = 1
                elif(tmp_size < vsiz_dec_last):
                    print('error-----'+str(k))

            if(vsz_value == ''):
                if(sta_value != '' and sta_next != ''):
                    tmp_size = int(sta_next,16) - int(sta_value,16)
                    vsz_value = tmp_size
                    vsize_dec = tmp_size
                    write_cell(wb_sheet,k,vsz_ind,conv_write_str(tmp_size,'size'),style)
                    if(ws_temp.cell(k,vsz_ind+1).value is None):
                        write_cell(wb_sheet,k,vsz_ind+1,'RSVD',style)
                elif(sta_value != '' and end_value != ''):
                    tmp_size = int(end_value,16) - int(sta_value,16) + 1
                    vsz_value = tmp_size
                    vsize_dec = tmp_size
                    write_cell(wb_sheet,k,vsz_ind,conv_write_str(tmp_size,'size'),style)
                    if(ws_temp.cell(k,vsz_ind+1).value is None):
                        write_cell(wb_sheet,k,vsz_ind+1,'RSVD',style)                
                else:
                    write_cell(wb_sheet,k,vsz_ind,'N/A!',style_warning)

            if(siz_value == ''):
                if(sta_value != '' and end_value != ''):
                    tmp_size = int(end_value,16) - int(sta_value,16) + 1
                    siz_value = tmp_size
                    size_dec = tmp_size
                    write_cell(wb_sheet,k,siz_ind,conv_write_str(tmp_size,'size'),style)
                elif(vsz_value != '' and vsize_dec > 0):
                    siz_value = vsz_value
                    size_dec = vsize_dec
                    tmp_size = vsize_dec
                    write_cell(wb_sheet,k,siz_ind,conv_write_str(tmp_size,'size'),style)                    
                else:
                    write_cell(wb_sheet,k,siz_ind,'N/A!',style_warning)

            if(end_value == ''):
                if(sta_value != '' and siz_value != ''):
                    tmp_end = int(sta_value,16)+size_dec-1
                    end_value = conv_write_str(tmp_end,'hex')
                    write_cell(wb_sheet,k,end_ind,end_value,style)
                else:
                    write_cell(wb_sheet,k,end_ind,'N/A!',style_warning)
                    
        cnt = 0
        wb2.save('fill.xlsx')#保存数据
        wb_fill = openpyxl.load_workbook('fill.xlsx',data_only = "true")
        ws_fill = wb_fill[sheets1[i]]
        
        for ll in range(num_rows):
            #print(ins_flag[ll])
            if(ins_flag[ll] == 1):

                if(ll==1):
                    print('error----'+str(ll))                
                else:

                    sta_last = ws_fill.cell(ll-1,sta_ind).value
                    vsz_last = ws_fill.cell(ll-1,vsz_ind).value
                                    
                    sta_value = ws_fill.cell(ll,sta_ind).value        
                    vsz_value = ws_fill.cell(ll,vsz_ind).value

                    
                    
                    if(vsz_last != ''):                        
                        vsz_dec_last = size2dec(vsz_last)
                    else:
                        vsz_dec_last = 0
                        
                    if(vsz_value != ''):
                        vsize_dec = size2dec(vsz_value)
                    else:
                        vsize_dec = 0

                    tmp_start = int(sta_last,16)+vsz_dec_last;
                    tmp_size = int(sta_value,16) - int(sta_last,16) - vsz_dec_last
                    tmp_end = tmp_start + tmp_size -1;

                    wb_sheet.insert_rows(ll+cnt)#插入行

                    write_cell(wb_sheet,ll+cnt,sta_ind,conv_write_str(tmp_start,'hex'),style_insert)
                    write_cell(wb_sheet,ll+cnt,end_ind,conv_write_str(tmp_end,'hex'),style_insert)
                    write_cell(wb_sheet,ll+cnt,vsz_ind+1,'RSVD',style_insert)
                    write_cell(wb_sheet,ll+cnt,vsz_ind,conv_write_str(tmp_size,'size'),style_insert)
                    write_cell(wb_sheet,ll+cnt,siz_ind,conv_write_str(tmp_size,'size'),style_insert)
                    
                    cnt = cnt + 1
    print('------------------')
    
    
wb2.save('result.xlsx')#保存数据
wb1.close()#关闭excel
wb2.close()
wb_fill.close()
os.remove('fill.xlsx')

##wb3=openpyxl.load_workbook('test.xlsx',data_only = "true")
##wsname=wb3.sheetnames#获取sheet页
##print(swname)
##nsheets = len(swname)
##ss = wb3.worksheets[0]

##import openpyxl
##
##wb = openpyxl.load_workbook("D:\村数据\实验.xlsx")
##ws = wb["Sheet1"]
##
##ws.insert_rows(3)#插入行
##ws.insert_cols(4)#插入列
##
##wb.save("D:\村数据\实验.xlsx")
##————————————————
##版权声明：本文为CSDN博主「农大丶冰封」的原创文章，遵循CC 4.0 BY-SA版权协议，转载请附上原文出处链接及本声明。
##原文链接：https://blog.csdn.net/qq_20075991/article/details/119448373



##ss.cell(1,1).value = 'aaaaaaaa'
##wb3.save('test.xlsx')
##wb3.close()
