import copy
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import glob
import numpy as np


def act_max_row(sheet_rm_blank,nlist):
    wm_rm = list(sheet_rm_blank.merged_cells)
    b_min = []
    b_max = []
    b_cmb = []
    mark_rm_rows = []
    num_rows = sheet_rm_blank.max_row
    if len(wm_rm) > 0:
        for i in range(0, len(wm_rm)):
            if(wm_rm[i].max_row > wm_rm[i].min_row):
                if(wm_rm[i].max_row not in b_cmb and wm_rm[i].min_row not in b_cmb):
                    b_cmb = b_cmb + list(range(wm_rm[i].min_row,wm_rm[i].max_row+1))
                elif(wm_rm[i].max_row in b_cmb and wm_rm[i].min_row not in b_cmb):
                    b_cmb = b_cmb + list(range(wm_rm[i].min_row,min(b_cmb)))                  
                elif(wm_rm[i].max_row not in b_cmb and wm_rm[i].min_row in b_cmb):
                    b_cmb = b_cmb + list(range(max(b_cmb)+1,wm_rm[i].max_row+1))

    for kk in range(num_rows):        
        if(sheet_rm_blank.cell(kk+1,nlist[0]).value is None and \
           sheet_rm_blank.cell(kk+1,nlist[1]).value is None and \
           sheet_rm_blank.cell(kk+1,nlist[2]).value is None and \
           sheet_rm_blank.cell(kk+1,nlist[-1]).value is None and \
           kk+1 not in b_cmb):
           mark_rm_rows.append(kk+1)

    return mark_rm_rows


##wb2 = openpyxl.load_workbook('./temp/AD1000_Master_Slave_Requirements_FSI.xlsx')
##sheetnames = ['Master', 'Slave']
##
##for sheet_rm_name in sheetnames:    
##    ss1 = wb2[sheet_rm_name]
##    print(sheet_rm_name)
##    print(act_max_row(ss1,[4,5,14,15]))
##
##wb2.close()
