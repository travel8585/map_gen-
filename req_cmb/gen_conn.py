import copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Color,colors,Border,Side,Alignment,PatternFill
import os
import time
import glob
import numpy as np
from extlib import act_max_row
from tkinter import filedialog
import sys

#user_path = os.path.dirname(__file__)
#folder_path=user_path+os.sep+time.strftime("%Y-%m-%d")
#folder_path = user_path+os.sep+'temp\\'
#folder_path = 'C:\\DEF\\siengine_svn\\1-Project_Development\\AD1000\\Dev\\2-Design\\Subsystem_Design\\NOC_SS\\1-spec\\2-ip_requirement\\M0\\'
folder_path = './'
#list_file=os.listdir(folder_path)
#list_file=glob.glob(folder_path+'*ments_*.xlsx')
list_file = filedialog.askopenfilenames(initialdir = folder_path)


#print(list_file)

if(len(list_file)==0):
    print('error : no files specified!')
    sys.exit(-1)
    
else:
    print('%d files were selected!\n' % len(list_file))
    

row_offset = [0,0]
save_path = 'conn.xlsx'
wb2 = openpyxl.Workbook()
sheet2 = wb2.create_sheet('connectivity')
for k in range(len(list_file)):

    path = list_file[k]
    wb = openpyxl.load_workbook(path)

    sheetname_wb =wb.sheetnames
    sheetnames = ['Master', 'Slave']    
    print('******'+list_file[k]+'********')
    print(sheetname_wb)
    for index,sheetname in enumerate(sheetnames):
        print(list_file[k]+'-----'+sheetname)
        if sheetname not in sheetname_wb :
            print('sheet not found!!')
        else:
            sheet = wb[sheetname]
            row_rm = act_max_row(sheet,[4,5,14,15])
            print(row_rm)
            wm = list(sheet.merged_cells)

            for i, row in enumerate(sheet.iter_rows()):

                if(i+1 in row_rm):
                    print('='*50)
                    print('\nWarning: row %d removed in sheet %s\n' %(i+1,sheetname))
                    print('='*50)
                    continue                
                for j, cell in enumerate(row):
                    if(j==0 and i>3):                        
                        if(sheetname == 'Master'):
                            sheet2.cell(row=i + 1 + row_offset[index]-2, column=j + 1, value=cell.value)
                        elif(sheetname == 'Slave'):
                            sheet2.cell(row=j+1, column=i + 1 + row_offset[index]-2, value=cell.value)
                    elif(j==4 and i>3 and sheetname == 'Master'):
                        sheet2.cell(row=i + 1 + row_offset[index]-2, column=j -2, value=cell.value)
                    elif(j==5 and i>3 and sheetname == 'Slave'):
                        sheet2.cell(row=j-3, column=i + 1 + row_offset[index]-2, value=cell.value)
                    else:
                        continue
                        
                        # 设置单元格格式
                        #source_cell = sheet.cell(i+1, j+1)
                        #target_cell = sheet2.cell(i+1+row_offset[index], j+1)
                        #target_cell.fill = copy.copy(source_cell.fill)
                        #if source_cell.has_style:
                            #target_cell._style = copy.copy(source_cell._style)
                            #target_cell.font = copy.copy(source_cell.font)
                            #target_cell.border = copy.copy(source_cell.border)
                            #target_cell.fill = copy.copy(source_cell.fill)
                            #target_cell.number_format = copy.copy(source_cell.number_format)
                            #target_cell.protection = copy.copy(source_cell.protection)
                            #target_cell.alignment = copy.copy(source_cell.alignment)
        row_offset[index]=row_offset[index]+sheet.max_row-4-len(row_rm)
        print(row_offset)


## set styles ###


border_set = Border(left=Side(style='thin', color=colors.BLACK),  # 左边框
                    right=Side(style='thin', color=colors.BLACK),  # 右边框
                    top=Side(style='thin', color=colors.BLACK),  # 上边框
                    bottom=Side(style='thin', color=colors.BLACK))  # 下边框

alignment1 = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False)
alignment2 = Alignment(horizontal='center', vertical='bottom', text_rotation=90, wrap_text=False)

font_bold = Font(bold = True)
fill1 = PatternFill('solid', fgColor="FFE4C4")
fill2 = PatternFill('solid', fgColor="90EE90")

## delete empty row and columns ###
del_c = sheet2.max_column
del_r = sheet2.max_row
print('original:  row = %d , column = %d' %(del_r,del_c))


del_num = 0

for kk in range(3,del_r+1):
    if(sheet2.cell(kk-del_num,2).value is None):        
        #print(kk)
        #print('x'*50)
        sheet2.delete_rows(kk-del_num)
        del_num = del_num + 1

del_num = 0

for kk in range(3,del_c+1):
    if(sheet2.cell(2,kk-del_num).value is None):
        #print(kk)
        #print('y'*50)
        sheet2.delete_cols(kk-del_num)
        del_num = del_num + 1
        

#str = sheet2.cell(2,kk).value
#if(str.startswith('scp_')):
#    print(sheet2.cell(2,kk).value)


## merge cells #####


del_c = sheet2.max_column
del_r = sheet2.max_row
print('after deleting:  row = %d , column = %d' %(del_r,del_c))


merge_cnt = 0
for kk in range(3,del_r+1):
    sheet2.cell(kk,1).font = font_bold
    sheet2.cell(kk,1).alignment = alignment1
    if(sheet2.cell(kk,1).value is not None):
        if(kk > 3 and merge_start < merge_end):            
            sheet2.merge_cells(start_row = merge_start,end_row = merge_end,start_column = 1,end_column = 1)
        merge_cnt = not merge_cnt            
        merge_start = kk        
    else:
        merge_end = kk
    if(merge_cnt == 0):
        sheet2.cell(kk,2).fill = fill1
    else:
        sheet2.cell(kk,2).fill = fill2        
    
if(merge_end == del_r and merge_end > merge_start):
    sheet2.merge_cells(start_row = merge_start,end_row = merge_end,start_column = 1,end_column = 1)


        
merge_cnt = 0
for kk in range(3,del_c+1):
    sheet2.cell(1,kk).font = font_bold
    sheet2.cell(1,kk).alignment = alignment1    
    if(sheet2.cell(1,kk).value is not None):
        if(kk > 3 and merge_start < merge_end):
            sheet2.merge_cells(start_row = 1,end_row = 1,start_column = merge_start,end_column = merge_end)
        merge_cnt = not merge_cnt
        merge_start = kk
    else:
        merge_end = kk
    if(merge_cnt == 0):
        sheet2.cell(2,kk).fill = fill1
    else:
        sheet2.cell(2,kk).fill = fill2
    sheet2.cell(2,kk).alignment = alignment2
    sheet2.column_dimensions[get_column_letter(kk)].width = 3
    
if(merge_end == del_c and merge_end > merge_start):
    sheet2.merge_cells(start_row = 1,end_row = 1,start_column = merge_start,end_column = merge_end)


for mm in range(1,del_r+1):
    for nn in range(1,del_c+1):
        sheet2.cell(mm,nn).border = border_set
    
        
if 'Sheet' in wb2.sheetnames:
    del wb2['Sheet']
    
wb.close()
wb2.save(save_path)
wb2.close()
print('Done.')
#os.open('conn.xlsx',os.O_RDONLY)
os.system(r'start conn.xlsx')
