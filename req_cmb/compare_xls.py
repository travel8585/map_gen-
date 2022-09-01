import copy
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import glob
import numpy as np
from extlib import act_max_row
from tkinter import filedialog
import sys

folder_path = 'C:\\DEF\\siengine_svn\\1-Project_Development\\AD1000\\Dev\\2-Design\\Subsystem_Design\\NOC_SS\\1-spec\\2-ip_requirement\\M0\\'
list_file0 = filedialog.askopenfilename(initialdir = folder_path)
list_file1 = filedialog.askopenfilename(initialdir = folder_path)

print(list_file0+'-----'+list_file1)


if(list_file0 =='' or list_file1 ==''):
    print('error : number of compared files should be 2')
    sys.exit(-1)
else:
    print('%d files were selected!\n' % (len(list_file0)+len(list_file1)))

save_path = 'cm_res.xlsx'
wb2 = openpyxl.Workbook()
wb0 = openpyxl.load_workbook(list_file0)
wb1 = openpyxl.load_workbook(list_file1)
    
sheetname_wb0 =wb0.sheetnames
sheetname_wb1 =wb0.sheetnames
sheetnames = ['Master', 'Slave']        
print(sheetname_wb0)
print('-------------')
print(sheetname_wb1)

diff_cnt = 0

for index,sheetname in enumerate(sheetnames):
    if sheetname not in sheetname_wb0 or sheetname not in sheetname_wb1 :
        print('sheet not found!!')
    else:
        sheet0 = wb0[sheetname]
        sheet1 = wb1[sheetname]
        sheet2 = wb2.create_sheet(sheetname)
        
        for i, row in enumerate(sheet0.iter_rows()):
            for j, cell in enumerate(row):
                if(sheet0.cell(i + 1,j + 1).value != sheet1.cell(i + 1,j + 1).value):
                    sheet2.cell(row=i + 1,column=j + 1).value = 'xxx'
                    diff_cnt = diff_cnt + 1
    
if 'Sheet' in wb2.sheetnames:
    del wb2['Sheet']

wb0.close()
wb1.close()
wb2.save(save_path)
wb2.close()
print('Done.')
print('%d different values were found!'%diff_cnt)
os.system(r'start '+save_path)
