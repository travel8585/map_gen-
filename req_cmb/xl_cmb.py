import copy
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import glob
import numpy as np
from extlib import act_max_row

user_path = os.path.dirname(__file__)
#folder_path=user_path+os.sep+time.strftime("%Y-%m-%d")
folder_path = user_path+os.sep+'temp\\'
#folder_path = 'C:\\DEF\\siengine_svn\\1-Project_Development\\AD1000\\Dev\\2-Design\\Subsystem_Design\\NOC_SS\\1-spec\\2-ip_requirement\\M0\\'
#list_file=os.listdir(folder_path)
list_file=glob.glob(folder_path+'*ments_*.xlsx')
if(len(list_file)==0):
    print('error : no files specified!')
row_offset = [0,0]
save_path = 'test.xlsx'
wb2 = openpyxl.Workbook()
for k in range(len(list_file)):
    #print(folder_path+list_file[k])
    #path = folder_path+list_file[k]    
    #print(path)
    #print(save_path)
    path = list_file[k]
    #save_path = "数据-复制.xlsx"

    wb = openpyxl.load_workbook(path)
    
    #sheetnames = wb.sheetnames
    sheetname_wb =wb.sheetnames
    sheetnames = ['Master', 'Slave']    
    print('******'+list_file[k]+'********')
    print(sheetname_wb)
    for index,sheetname in enumerate(sheetnames):
        print(list_file[k]+'-----'+sheetname)
        if sheetname not in sheetname_wb :
            print('sheet not found!!')
        else:
            sheet = wb[sheetname]
            row_rm = act_max_row(sheet,[4,5,14,15])
            print(row_rm)
            if(k == 0):
                sheet2 = wb2.create_sheet(sheetname)
            else:
                sheet2 = wb2[sheetname]

            # tab颜色
            sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

            # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
            wm = list(sheet.merged_cells)
    ##        print(wm[0])
    ##        print(dir(wm[0]))
    ##        print(wm[0].max_row)
    ##        print(wm[0].max_col)
    ##        print(wm[0].min_row)
    ##        print(wm[0].min_col)
    ##        print(wm[0].left)
    ##        print(wm[0].right)
    ##        print(wm[0].size)
            if len(wm) > 0:
                for i in range(0, len(wm)):                
                    #print(str(wm[i]))
                    #cell2 = str(wm[i])
                    cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')            
                    #print(cell2)
                    #sheet2.merge_cells(cell2)
                    if((k == 0 or wm[i].min_row > 3) and wm[i].min_row not in row_rm and wm[i].max_row not in row_rm ):
                        sheet2.merge_cells(start_row=wm[i].min_row+row_offset[index], end_row=wm[i].max_row+row_offset[index], start_column=wm[i].min_col, end_column=wm[i].max_col)
            for i, row in enumerate(sheet.iter_rows()):
                #print(i)
                #print(row)
                if(i+1 in row_rm):
                    #print('row %d removed in sheet %s' %i+1 sheetname)
                    print('='*50)
                    print('\nWarning: row %d removed in sheet %s\n' %(i+1,sheetname))
                    print('='*50)
                    continue
                sheet2.row_dimensions[i+1+row_offset[index]].height = sheet.row_dimensions[i+1].height                
                for j, cell in enumerate(row):
                    if(k==0 or i>3):
                        sheet2.column_dimensions[get_column_letter(j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
                        sheet2.cell(row=i + 1 + row_offset[index], column=j + 1, value=cell.value)
                        
                        # 设置单元格格式
                        source_cell = sheet.cell(i+1, j+1)
                        target_cell = sheet2.cell(i+1+row_offset[index], j+1)
                        target_cell.fill = copy.copy(source_cell.fill)
                        if source_cell.has_style:
                            target_cell._style = copy.copy(source_cell._style)
                            target_cell.font = copy.copy(source_cell.font)
                            target_cell.border = copy.copy(source_cell.border)
                            target_cell.fill = copy.copy(source_cell.fill)
                            target_cell.number_format = copy.copy(source_cell.number_format)
                            target_cell.protection = copy.copy(source_cell.protection)
                            target_cell.alignment = copy.copy(source_cell.alignment)
        row_offset[index]=row_offset[index]+sheet.max_row-4-len(row_rm)
        print(row_offset)
        
if 'Sheet' in wb2.sheetnames:
    del wb2['Sheet']
    
wb.close()
wb2.save(save_path)
wb2.close()
print('Done.')
